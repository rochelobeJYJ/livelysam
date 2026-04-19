#!/usr/bin/env python
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import requests
from google.auth.transport.requests import Request
from google.oauth2 import service_account
from openpyxl import load_workbook


SCOPES = ["https://www.googleapis.com/auth/datastore"]
FIRESTORE_BASE_URL = "https://firestore.googleapis.com/v1"
NAME_KEYS = ("nickname", "name", "player", "username", "닉네임", "이름")
SCORE_KEYS = ("score", "highscore", "record", "점수", "기록")
GAME_KEYS = ("gameid", "game_id", "game", "게임", "mode", "sheet")
RANK_KEYS = ("rank", "순위")
DEFAULT_SOURCE = "legacy-import"
SHEET_NAME_ALIASES = {
    "dino run": "dino-run-1",
    "dino run 1": "dino-run-1",
    "dinorun": "dino-run-1",
    "dinorun1": "dino-run-1",
    "dino1": "dino-run-1",
    "dino run 2": "dino-run-2",
    "dinorun2": "dino-run-2",
    "dino2": "dino-run-2",
    "dino run hard": "dino-run-hard",
    "dino hard": "dino-run-hard",
    "hard": "dino-run-hard",
    "wing tap 1": "wing-tap-1",
    "wing tap classic": "wing-tap-1",
    "wingtap1": "wing-tap-1",
    "game1": "wing-tap-1",
    "wing tap 2": "wing-tap-2",
    "wingtap2": "wing-tap-2",
    "game2": "wing-tap-2",
    "wing boss": "wing-boss",
    "wingboss": "wing-boss",
    "game3": "wing-boss",
}


@dataclass
class LeaderboardRow:
    game_id: str
    nickname: str
    score: int
    rank_hint: str
    sheet_name: str


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def normalize_header(value: object) -> str:
    return normalize_text(value).replace(" ", "").replace("-", "").replace("_", "").lower()


def normalize_sheet_name(value: str) -> str:
    return " ".join(normalize_text(value).lower().replace("_", " ").split())


def normalize_game_id(value: str) -> str:
    text = normalize_sheet_name(value)
    return SHEET_NAME_ALIASES.get(text, text.replace(" ", "-"))


def pick_value(row: dict[str, object], keys: Iterable[str]) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = normalize_text(value)
        if text:
            return text
    return ""


def parse_score(value: object) -> int | None:
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def build_document_id(row: LeaderboardRow) -> str:
    seed = f"{row.game_id}|{row.nickname}|{row.score}|{row.rank_hint}|{row.sheet_name}"
    return f"legacy-{hashlib.sha1(seed.encode('utf-8')).hexdigest()[:20]}"


def build_firestore_document(row: LeaderboardRow, season_id: str, source: str, timestamp: str) -> dict[str, object]:
    player_id = build_document_id(row)
    return {
        "player_id": player_id,
        "document": {
            "fields": {
                "seasonId": {"stringValue": season_id},
                "gameId": {"stringValue": row.game_id},
                "playerId": {"stringValue": player_id},
                "nickname": {"stringValue": row.nickname[:12]},
                "score": {"integerValue": str(max(0, row.score))},
                "source": {"stringValue": source},
                "createdAt": {"timestampValue": timestamp},
                "updatedAt": {"timestampValue": timestamp},
                "lastSubmittedAt": {"timestampValue": timestamp},
            }
        },
    }


def build_token(credentials_path: Path) -> str:
    credentials = service_account.Credentials.from_service_account_file(
        str(credentials_path),
        scopes=SCOPES,
    )
    credentials.refresh(Request())
    return credentials.token


def iter_csv_rows(path: Path, forced_game_id: str | None) -> Iterable[LeaderboardRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return
        header_map = {name: normalize_header(name) for name in reader.fieldnames}
        for index, raw_row in enumerate(reader, start=2):
            row = {header_map[key]: value for key, value in raw_row.items() if key is not None}
            parsed = parse_row(row, forced_game_id=forced_game_id, sheet_name=path.stem, fallback_rank=index)
            if parsed:
                yield parsed


def iter_xlsx_rows(path: Path, forced_game_id: str | None) -> Iterable[LeaderboardRow]:
    workbook = load_workbook(path, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = values[0]
            header_map = {index: normalize_header(cell) for index, cell in enumerate(headers)}
            for row_index, values_row in enumerate(values[1:], start=2):
                row = {
                    header_map[index]: values_row[index]
                    for index in range(len(headers))
                    if header_map.get(index)
                }
                parsed = parse_row(
                    row,
                    forced_game_id=forced_game_id,
                    sheet_name=worksheet.title,
                    fallback_rank=row_index,
                )
                if parsed:
                    yield parsed
    finally:
        workbook.close()


def parse_row(row: dict[str, object], forced_game_id: str | None, sheet_name: str, fallback_rank: int) -> LeaderboardRow | None:
    nickname = pick_value(row, NAME_KEYS)
    score_value = pick_value(row, SCORE_KEYS)
    game_id = forced_game_id or pick_value(row, GAME_KEYS) or normalize_game_id(sheet_name)
    rank_hint = pick_value(row, RANK_KEYS) or str(fallback_rank)
    score = parse_score(score_value)

    if not nickname or score is None:
        return None

    return LeaderboardRow(
        game_id=game_id,
        nickname=nickname,
        score=score,
        rank_hint=rank_hint,
        sheet_name=sheet_name,
    )


def load_rows(path: Path, forced_game_id: str | None) -> list[LeaderboardRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return list(iter_csv_rows(path, forced_game_id))
    if suffix == ".xlsx":
        return list(iter_xlsx_rows(path, forced_game_id))
    raise ValueError(f"Unsupported input type: {path.suffix}")


def write_rows(
    rows: list[LeaderboardRow],
    *,
    project_id: str,
    season_id: str,
    credentials_path: Path,
    source: str,
    dry_run: bool,
) -> None:
    timestamp = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    documents = [build_firestore_document(row, season_id, source, timestamp) for row in rows]
    print(f"Loaded {len(documents)} rows")

    if dry_run:
        for item in documents[:10]:
            print(json.dumps(item, ensure_ascii=False))
        if len(documents) > 10:
            print(f"... and {len(documents) - 10} more rows")
        return

    token = build_token(credentials_path)
    session = requests.Session()
    session.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })

    success = 0
    for item in documents:
        row = item["document"]["fields"]
        game_id = row["gameId"]["stringValue"]
        document_id = item["player_id"]
        url = (
            f"{FIRESTORE_BASE_URL}/projects/{project_id}/databases/(default)/documents/"
            f"minigameLeaderboards/{season_id}/games/{game_id}/entries/{document_id}"
        )
        response = session.patch(url, json=item["document"], timeout=30)
        if response.status_code >= 300:
            print(f"[ERROR] {game_id}/{document_id}: {response.status_code} {response.text}", file=sys.stderr)
            response.raise_for_status()
        success += 1

    print(f"Imported {success} rows into project '{project_id}' season '{season_id}'")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import legacy minigame leaderboard rows from CSV/XLSX into Firestore."
    )
    parser.add_argument("--input", required=True, help="Path to .csv or .xlsx export file")
    parser.add_argument("--project-id", required=True, help="Firebase project id")
    parser.add_argument("--season-id", default="season-1", help="Firestore season id")
    parser.add_argument("--credentials", required=True, help="Service account json path")
    parser.add_argument("--game-id", help="Force a single game id for CSV or mixed exports")
    parser.add_argument("--source", default=DEFAULT_SOURCE, help="Stored source label")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, do not upload")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).resolve()
    credentials_path = Path(args.credentials).resolve()

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1
    if not credentials_path.exists():
        print(f"Credentials file not found: {credentials_path}", file=sys.stderr)
        return 1

    rows = load_rows(input_path, normalize_text(args.game_id) or None)
    if not rows:
        print("No valid rows found. Check headers or sheet names.", file=sys.stderr)
        return 1

    write_rows(
        rows,
        project_id=normalize_text(args.project_id),
        season_id=normalize_text(args.season_id) or "season-1",
        credentials_path=credentials_path,
        source=normalize_text(args.source) or DEFAULT_SOURCE,
        dry_run=bool(args.dry_run),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
